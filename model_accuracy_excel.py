import os
import pandas as pd
import numpy as np
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from copy import copy

ROOT = os.getcwd()

ACTUAL_XLSX = os.path.join(ROOT, "data", "fact_monthly_closed.xlsx")
ACTUAL_CSV  = os.path.join(ROOT, "data", "fact_monthly_closed.csv")

HISTORY_PATH = os.path.join(
    ROOT, "backend", "data", "logs", "forecast_horizon_history.csv"
)

OUT_PATH = os.path.join(
    ROOT, "backend", "data", "outputs", "model_accuracy_analysis.xlsx"
)

ACTUAL_COL = "Secondary_Sales_Qty"


def norm_code(s):
    return s.astype(str).str.strip().str.replace(r"\.0$", "", regex=True)


def month_label_from_actual(df):
    if "Month" in df.columns:
        m = pd.to_datetime(df["Month"], errors="coerce")
        if m.notna().any():
            return m.dt.strftime("%Y-%m")

    month_col = "Month_Number" if "Month_Number" in df.columns else "MonthNo"
    return (
        df["Year"].astype(int).astype(str)
        + "-"
        + df[month_col].astype(int).astype(str).str.zfill(2)
    )


def pretty_month(m):
    return pd.to_datetime(m + "-01").strftime("%b-%y")


def calc_accuracy(actual, forecast):
    actual = float(actual)
    forecast = float(forecast)

    if actual == 0 and forecast == 0:
        return 1.0
    if actual == 0 and forecast != 0:
        return 0.0

    acc = 1 - abs(actual - forecast) / actual
    return max(0.0, acc)


# -----------------------------
# Load actuals
# -----------------------------
if os.path.exists(ACTUAL_XLSX):
    actual = pd.read_excel(ACTUAL_XLSX)
elif os.path.exists(ACTUAL_CSV):
    actual = pd.read_csv(ACTUAL_CSV)
else:
    raise FileNotFoundError("fact_monthly_closed.xlsx/csv not found in data/")

actual["ItemCode"] = norm_code(actual["ItemCode"])
actual["Month_Label"] = month_label_from_actual(actual)
actual[ACTUAL_COL] = pd.to_numeric(actual[ACTUAL_COL], errors="coerce").fillna(0)

actual_m = (
    actual.groupby(["ItemCode", "Month_Label"], as_index=False)[ACTUAL_COL]
    .sum()
    .rename(columns={ACTUAL_COL: "Actual"})
)

# -----------------------------
# Load model forecast history only
# -----------------------------
hist = pd.read_csv(HISTORY_PATH)
hist["ItemCode"] = norm_code(hist["ItemCode"])

hist = hist[
    (hist["Horizon"].astype(str) == "M+1") &
    (hist["Forecast_Source"].astype(str) == "AI_CHAMPION_MODEL")
].copy()

hist["Forecast_Qty"] = pd.to_numeric(hist["Forecast_Qty"], errors="coerce").fillna(0)

if "Run_Date" in hist.columns:
    hist["Run_Date_sort"] = pd.to_datetime(hist["Run_Date"], errors="coerce")
    hist = hist.sort_values(["ItemCode", "Forecast_Month", "Run_Date_sort"])
else:
    hist = hist.sort_values(["ItemCode", "Forecast_Month"])

hist = hist.drop_duplicates(
    subset=["ItemCode", "Forecast_Month"],
    keep="last"
)

forecast_m = hist[["ItemCode", "Forecast_Month", "Forecast_Qty"]].rename(
    columns={
        "Forecast_Month": "Month_Label",
        "Forecast_Qty": "Forecast",
    }
)

# -----------------------------
# Merge actual + forecast
# -----------------------------
merged = forecast_m.merge(
    actual_m,
    on=["ItemCode", "Month_Label"],
    how="left"
)

merged["Actual"] = pd.to_numeric(merged["Actual"], errors="coerce")
merged = merged.dropna(subset=["Actual"]).copy()

merged["Accuracy"] = merged.apply(
    lambda r: calc_accuracy(r["Actual"], r["Forecast"]),
    axis=1
)

months = sorted(merged["Month_Label"].unique())

# -----------------------------
# Build wide Excel format
# -----------------------------
skus = sorted(merged["ItemCode"].unique())
out = pd.DataFrame({"SKU": skus})

for m in months:
    mm = pretty_month(m)
    temp = merged[merged["Month_Label"] == m].set_index("ItemCode")

    out[(mm, "Actual")] = out["SKU"].map(temp["Actual"])
    out[(mm, "Forecast")] = out["SKU"].map(temp["Forecast"])
    out[(mm, "Accuracy")] = out["SKU"].map(temp["Accuracy"])

# Multi header
cols = [("", "SKU")] + [c for c in out.columns if c != "SKU"]
out = out[["SKU"] + [c for c in out.columns if c != "SKU"]]

new_cols = [("", "SKU")]
for c in out.columns[1:]:
    new_cols.append(c)

out.columns = pd.MultiIndex.from_tuples(new_cols)

os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)

with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:

    out.to_excel(writer, index=True, sheet_name="Model Accuracy")

    ws = writer.book["Model Accuracy"]

    # remove pandas index column

    ws.delete_cols(1)

    # formatting

    for row in ws.iter_rows():

        for cell in row:

            cell.alignment = copy(cell.alignment)

            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "B3"

    for col_idx in range(1, ws.max_column + 1):

        col_letter = get_column_letter(col_idx)

        max_len = 0

        for row_idx in range(1, ws.max_row + 1):

            value = ws.cell(row=row_idx, column=col_idx).value

            max_len = max(max_len, len(str(value or "")))

        ws.column_dimensions[col_letter].width = max(12, min(max_len + 2, 18))

    # percentage format for Accuracy columns

    for col_idx in range(1, ws.max_column + 1):

        header2 = ws.cell(row=2, column=col_idx).value

        if header2 == "Accuracy":

            for row_idx in range(3, ws.max_row + 1):

                ws.cell(row=row_idx, column=col_idx).number_format = "0%"


print("Saved:", OUT_PATH)
print("Rows:", len(out))
print("Months:", [pretty_month(m) for m in months])